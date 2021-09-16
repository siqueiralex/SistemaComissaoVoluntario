import xlwt
import string
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment
from openpyxl.styles import PatternFill

class NewSheetBuilder:
    _curr_rand_color = 0
    _colors = [PatternFill(fgColor='E5FFCC', fill_type = "solid"), \
               PatternFill(fgColor='FFE5CC', fill_type = "solid"), \
               PatternFill(fgColor='CCFFFF', fill_type = "solid"), \
               PatternFill(fgColor='E0E0E0', fill_type = "solid"), \
               PatternFill(fgColor='CCCCFF', fill_type = "solid"), \
               PatternFill(fgColor='E6A4D9', fill_type = "solid"), \
               PatternFill(fgColor='7CB3DC', fill_type = "solid"), \
               PatternFill(fgColor='FF5733', fill_type = "solid"), \
               PatternFill(fgColor='E6C873', fill_type = "solid"), \
               PatternFill(fgColor='8AB5A1', fill_type = "solid")]
    _yellow_fill =  PatternFill(fgColor=colors.YELLOW, fill_type = "solid")
    _orange_fill =  PatternFill(fgColor='FFA500', fill_type = "solid")
    num2letter = dict(enumerate(string.ascii_uppercase, 1))
    _thin = Side(style='thin', color="000000")
    _thin_border = Border(left=_thin, top=_thin, right=_thin, bottom=_thin)
    _font_bold_big =  Font(bold=True, size=12)
    _font_bold =  Font(bold=True)
    _center_wrap = Alignment(horizontal='center',vertical='center',text_rotation=0,wrap_text=True,shrink_to_fit=False,indent=0)
    _center = Alignment(horizontal='center',vertical='center',text_rotation=0,wrap_text=False,shrink_to_fit=True,indent=0)
    _left_wrap = Alignment(horizontal='left',vertical='center',text_rotation=0,wrap_text=True,shrink_to_fit=False,indent=0)
    _left = Alignment(horizontal='left',vertical='center',text_rotation=0,wrap_text=False,shrink_to_fit=True,indent=0)
    
    styles = [NamedStyle(name="default", border = _thin_border, alignment=_center),
            NamedStyle(name="header", border = _thin_border, fill=_orange_fill, font=_font_bold_big, alignment=_center_wrap),
            NamedStyle(name="highlight_", border = _thin_border, fill=_yellow_fill, alignment=_center),
            NamedStyle(name="center_currency", border= _thin_border, alignment=_center, number_format="#,##0.00R$" )   
             ]


    def __init__(self, sheet):    
        self.sheet = sheet
        self.col=1
        self.row=1
        self.col_style = {}
        for style in self.styles:
            if style.name not in sheet.parent.style_names:
                sheet.parent.add_named_style(style)
        
    def enter(self):
        self.col=1
        self.row+=1
        
    def set_col_widths(self, widths):
        for i in range(1,len(widths)+1):
            for i, column_width in enumerate(widths, start=1):
                self.sheet.column_dimensions[self.num2letter[i]].width = column_width
    
    def get_random_color_style(self, seed=0):
        
        color_index = self._curr_rand_color % len(self._colors)
        self._curr_rand_color+=1
        if seed > 0:
            color_index = seed % len(self._colors)
            self._curr_rand_color=seed+1
        name = f"random_{color_index}"
        if name not in self.sheet.parent.style_names:
            self.sheet.parent.add_named_style(NamedStyle(name=name, border = self._thin_border, fill=self._colors[color_index], alignment=self._center))
        
        
        return name
        
        
    def write_cell(self, value, style=None, carr_ret=False):
        _style = "default"
        if style:
            _style = style
        if self.col in self.col_style:
            _style = self.col_style[self.col]
            
        cell = self.sheet.cell(row=self.row, column=self.col, value=value)
        cell.style = _style
        
        self.col+=1
        if carr_ret:
            self.enter()
        
    
    def write_line(self, line):
        for cell in line:
            self.write_cell(cell)
        
    def write_header(self, header, style="header"):
        for item in header:
            if 'style' in item:
                self.col_style[self.col] = item['style']

            cell = self.sheet.cell(row=self.row, column=self.col, value=item['text'])
            cell.style = style
            self.col +=1


class SheetBuilder:

    def bg_s(color):
        return f"pattern: pattern solid, pattern_fore_colour {color};"

    center_s = "align: horiz center, vert center;"
    wrap_s = "align: wrap on;"
    border_s = "borders: top thin, bottom thin, left thin,right thin;"
    bold_s = "font: bold 1, color black;"
    date_format = 'dd/mm/yyyy'
    currency_format = 'R$#,##0.00'
    n2l = dict(enumerate(string.ascii_uppercase, 0))    

    style_dict = {
        'header' : xlwt.easyxf(center_s+wrap_s+border_s+bold_s+bg_s("light_orange")),
        'highlight' : xlwt.easyxf(center_s+border_s+bg_s("light_yellow")),
        'left' : xlwt.easyxf(border_s),
        'left160' : xlwt.easyxf(border_s+"font:height 160;"),
        'center' : xlwt.easyxf(center_s+border_s),
        'center_wrap': xlwt.easyxf(center_s+border_s+wrap_s),
        'center160' : xlwt.easyxf(center_s+border_s+"font:height 160;"),
        'center_currency' : xlwt.easyxf(center_s+border_s, num_format_str=currency_format),
        'center_currency_dark' : xlwt.easyxf(center_s+border_s+bold_s+bg_s("gray25"), num_format_str=currency_format),
        'center_currency_darker' : xlwt.easyxf(center_s+border_s+bold_s+bg_s("gray40"), num_format_str=currency_format),
        'center_date' : xlwt.easyxf(center_s+border_s, num_format_str=date_format),
        'total' : xlwt.easyxf(center_s+border_s+bold_s+bg_s("gray25")),
        'total_darker' : xlwt.easyxf(center_s+border_s+bold_s+bg_s("gray40")),
    }

    def num2letter(self, num):
        if num < 26:
            return self.n2l[num]
        return "A"+self.n2l[num%26]    

    def __init__(self, sheet):    
        self.sheet = sheet
        self.style = "center"
        self.row = 0
        self.col = 0
        self.col_initial = 0

        self.col_style = {}
    

    

    
    def get_style(self,):
        return self.style_dict[self.style]
    
    def set_col_widths(self, widths):
        for i in range(len(widths)):
            self.sheet.col(self.col_initial + i).width=widths[i]
    
    def write_header(self, header):
        self.col = self.col_initial
        for item in header:
            if 'style' in item:
                self.col_style[self.col] = item['style']

            self.sheet.write(self.row, self.col, item['text'], self.get_style())
            self.col +=1
    
    def enter(self):
        self.row+=1
        self.col=self.col_initial
    
    def write_line(self, line):
        for cell in line:
            self.write_col(cell)

    def write_col(self, value, **kwargs):
        if "style" in kwargs:
            style = kwargs['style']
        else:    
            style = self.get_style()
            if self.col in self.col_style:
                style = self.style_dict[self.col_style[self.col]]
        
        self.sheet.write(self.row,self.col,value,style)
        self.col+=1


    def write_sum(self, elig_col=[],  num_rows=0, rows=[]):
        for col in elig_col:
            col += self.col_initial
            kwargs = {}
            
            
            style = self.get_style()
            if col in self.col_style:
                if self.col_style[col] == "center_currency" and self.style == "total":
                    style = self.style_dict['center_currency_dark']
                elif self.col_style[col] == "center_currency" and self.style == "total_darker":
                    style = self.style_dict['center_currency_darker']
                else:
                    style =  self.style_dict[self.col_style[col]]
            
            formula = ""
            if num_rows > 0:
                formula = xlwt.Formula(f"SUM({self.num2letter(col)}{self.row - num_rows}:{self.num2letter(col)}{self.row})")
            if rows:
                formula = xlwt.Formula("+".join([f"{self.num2letter(col)}{i+1}" for i in rows]))
            
            self.sheet.write(self.row,col,formula,style)
        
        self.col = max(elig_col) + 1 + self.col_initial
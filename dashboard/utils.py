import xlwt
from openpyxl import Workbook

from django.contrib.auth.decorators import login_required
from django.utils.formats import date_format
import datetime


from .sheet_builder import SheetBuilder, NewSheetBuilder
from projetos.models import Atividade, Cronograma, Unidade, Vaga
from projetos.decorators import *

CUSTO_COTA = 600


def style_random_color(number):
    color_list = ["ocean_blue", "light_green",  "light_turquoise", "lavender", "ice_blue", "gray25", "sea_green", "aqua"]
    return xlwt.easyxf(f"borders: top thin, bottom thin, left thin,right thin;font: bold 1, color black;align: horiz center, vert center;pattern: pattern solid, pattern_fore_colour {color_list[number % len(color_list)]};")
     

def controlePlanosSheetFiller(sheet, cronograma, unidades):
    total_rows = []
    sb = SheetBuilder(sheet)
    sb.row = 1
    sb.col_initial = 0
    sb.sheet.row(1).height_mismatch = True
    sb.sheet.row(1).height = 600

    sb.set_col_widths([5000,4000,7500,3000,2500,2000,2500,2500,6000])
    

    sb.style="header"
    sb.write_header([
                    {'text':'UNIDADE'},
                    {'text':'TIPO DE ATIVIDADE'},
                    {'text':'ATIVIDADE'},
                    {'text':'HORÁRIO'},
                    {'text':'CH'},
                    {'text':'QTDE DIAS'},
                    {'text':'QTDE DE SV'},
                    {'text':'QTDE COTAS'},
                    {'text':'CUSTO','style':"center_currency"}
                    ])
    
    for unidade in unidades:
        sb.style="center"
        atividades = Atividade.objects.filter(cronograma_id=cronograma.id, unidade_id=unidade.id).order_by('id')
        sb.enter()

        qtde = atividades.count()
        sb.sheet.write_merge(sb.row, sb.row+qtde, sb.col, sb.col, unidade.sigla, style_random_color(unidade.id) )
        
        sb.col+=1
        sb.col_initial+=1
        for atv in atividades:
            sb.write_line([atv.tipo_atividade, atv.nome, atv.horario, atv.carga_horaria, atv.quantidade_de_dias, atv.total_voluntarios, atv.total_cotas, atv.total_cotas*CUSTO_COTA])
            sb.enter()

        sb.style = "total"
        total_rows += [sb.row]
        sb.sheet.write_merge(sb.row,sb.row,sb.col,sb.col+3, "TOTAL",sb.get_style())
        sb.write_sum(elig_col=[4,5,6,7], num_rows=qtde-1)
        
        sb.col_initial-=1
        sb.enter()

    sb.enter()
    sb.style = "total_darker"
    sb.sheet.write_merge(sb.row,sb.row,sb.col,sb.col+4, "TOTAL GERAL", sb.get_style())
    sb.write_sum(elig_col=[5,6,7,8], rows = total_rows)
    



def cotachSheetFiller(sheet, cronograma, unidades):
    sb = SheetBuilder(sheet)
    sb.row = 1
    sb.col_initial = 0
    sb.sheet.row(1).height_mismatch = True
    sb.sheet.row(1).height = 600

    total_rows = []
    sb.set_col_widths([3000,8000]+[1200]*len(cronograma.lista_de_datas)+[3000,5000])

    sb.style="header"
    sb.write_header([
                    {'text':'UNIDADE'},
                    {'text':'ATIVIDADE','style':"center160"},
                    ] + [{'text':f"{d.day}"} for d in cronograma.lista_de_datas] + [
                    {'text':'CH (Horas)'},
                    {'text':'HORÁRIOS'},
                    ])

    for unidade in unidades:
        sb.style="center"
        atividades = Atividade.objects.filter(cronograma_id=cronograma.id, unidade_id=unidade.id).order_by('id')
        sb.enter()
        
        qtde = atividades.count()
        
        sb.sheet.write_merge(sb.row, sb.row+qtde, sb.col, sb.col, unidade.sigla, style_random_color(unidade.id))
        
        sb.col+=1
        sb.col_initial+=1
        for atv in atividades:
            sb.write_line([atv.nome] + atv.lista_efetivos + [atv.carga_horaria, atv.horario])
            sb.enter()

        sb.style = "total"
        total_rows += [sb.row]
        sb.sheet.write_merge(sb.row,sb.row,sb.col,sb.col, "TOTAL",sb.get_style())
        date_cols = list(range(1,len(cronograma.lista_de_datas)+1))
        
        sb.write_sum(elig_col=date_cols, num_rows=qtde-1)
        sb.write_line(["",""])
        
        sb.col_initial-=1

        sb.enter()
    
    sb.enter()
    sb.style = "total_darker"
    sb.sheet.write_merge(sb.row,sb.row,sb.col,sb.col+1, "TOTAL GERAL", sb.get_style())
    
    date_cols = list(range(2,len(cronograma.lista_de_datas)+2))
    sb.write_sum(elig_col=date_cols, rows = total_rows)
    sb.write_line(["",""])



def projetoSheetFiller(sheet, cronograma, unidade):
    atividades = Atividade.objects.filter(unidade_id= unidade.id, cronograma_id=cronograma.id).order_by('id')
    sb = SheetBuilder(sheet)

    sb.sheet.row(1).height_mismatch = True
    sb.sheet.row(1).height = 900
    sb.col_initial = 1
    sb.row = 1
    sb.set_col_widths([3000,5000,10000,3000,4000,3000,3000,3000,3000,3000,4000])

    sb.style = "header"
    sb.write_header([{'text':'UNIDADE'},
                     {'text':'TIPO DE PROJETO'},
                     {'text':'NOME DO PROJETO','style':"center160"},
                     {'text':'MÊS'},
                     {'text':'DATA','style':"center_date"},
                     {'text':'QTDE DE DIAS'},
                     {'text':'Nº SV/DIA'},
                     {'text':'Nº DE JOVENS ATENDIDOS'},
                     {'text':'Nº MÁX JOVENS SIMULT'},
                     {'text':'CH'},
                     {'text':'HORARIO DA ATIVIDADE'}
                     ])
   
    sb.style = "center"

    for atv in atividades:
        sb.enter()
        sb.write_line([atv.unidade.sigla, atv.tipo_atividade, atv.nome, f"{date_format(atv.cronograma.inicio_servicos, 'F')}",atv.cronograma.nome_curto, atv.quantidade_de_dias, atv.voluntarios_dia, atv.numero_jovens, atv.numero_jovens_sim ,atv.carga_horaria, atv.horario])

    sb.col_initial = 2
    sb.style = "header"
    sb.enter()
    sb.enter()
    sb.enter()
    sb.write_header([{'text':'ATIVIDADE','style':"center_wrap"},
                     {'text':'OBJETIVO DA UNIDADE','style':"center_wrap"}])

    
    for atv in atividades:
        sb.enter()
        sb.write_line([atv.nome, atv.objetivo])



def extract_list_from_sheet(sheet, header_row=1):
    row_count = sheet.max_row
    col_count = sheet.max_column
    header = [cell.value for cell in sheet[header_row]]

    values_list = []
    for row in sheet.iter_rows(min_row=header_row+1):
        values = {}
        for key, cell in zip(header, row):
            values[key] = cell.value
        
        if values['Matrícula']==None:
            break

        values_list.append(values)
    return header, values_list
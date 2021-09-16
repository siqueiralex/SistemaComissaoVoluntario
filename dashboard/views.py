from django.shortcuts import render
import xlwt
from django.utils.formats import date_format
import os
from django.http import HttpResponse
from django.shortcuts import render
from django.contrib import messages
from io import BytesIO
import openpyxl
from openpyxl.writer.excel import save_virtual_workbook
from django import forms
from datetime import datetime
import zipfile
from .sheet_builder import SheetBuilder, NewSheetBuilder
from projetos.models import Vaga, Unidade, Cronograma
from copy import deepcopy as copy
from django.contrib.auth.decorators import login_required
from .forms import DistribuicaoForm, AtestoImpedimentoForm, HomologacaoForm
from .utils import extract_list_from_sheet, controlePlanosSheetFiller, cotachSheetFiller

def index(request):
    return render(request, "dashboard/home.html")


def distribuicao(request):
    trimestres = {
        0: ["Janeiro", "Fevereiro", "Março"],
        1: ["Abril", "Maio", "Junho"],
        2: ["Julho", "Agosto", "Setembro"],
        3: ["Outubro", "Novembro", "Dezembro"]
    }

    form = DistribuicaoForm()
    if request.method == 'POST':
        form = DistribuicaoForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data.get('file')
            dia = form.cleaned_data.get('dia')
            linha_header = form.cleaned_data.get('linha_header')           
            nome_aba = form.cleaned_data.get('nome_aba')

            trimestre = trimestres[int((dia.month-1)/3)]
            
            book = openpyxl.load_workbook(filename=BytesIO(file.read()))   


            if not nome_aba in book.sheetnames:
                messages.warning(request,f"Não tem Aba com nome {nome_aba}")
                return render(request, 'dashboard/distribuicao.html', {'form': form})

            header, classif = extract_list_from_sheet(book[nome_aba], header_row=linha_header)
            
            
            required_in_header = ['IMPEDIMENTO','CH MÊS','CH TRIMESTRE','Nº','Nome','Matrícula','Unidade Lot.',\
                'Jornada de trab.','Nº plantão', 'Unidade 1', 'Unidade 2', 'Interesse em outra unidade']
            if not all(x in header for x in required_in_header):
                messages.warning(request,f"Houve um problema com o Cabeçalho da sua planilha, verifique se as informações estão corretas.\
                    \nObs: o cabeçalho da Classificação deve conter EXATAMENTE estes textos: "+", ".join(required_in_header))
                return render(request, 'dashboard/distribuicao.html', {'form': form})
            inscritos = copy(classif)


            vagas = Vaga.objects.filter(dia=dia).order_by('-carga_horaria').values("dia","unidade__sigla", "atividade", "responsavel","horario", "carga_horaria")
            vagas = list(vagas)


            vagas_rem = copy(vagas)

            espera = []
            atribuicoes = {}
            for inscrito in inscritos:

                preferencias = (inscrito['Unidade 1'], inscrito['Unidade 2'])
                restrito = inscrito['Interesse em outra unidade'] == 'NÃO'

                unidade_vaga = {}
                for vaga in vagas_rem:
                    if vaga['unidade__sigla'] in preferencias or not restrito:
                        if not (vaga['carga_horaria'] + inscrito['CH MÊS']) > 48:
                            if not vaga['unidade__sigla'] in unidade_vaga:
                                unidade_vaga[vaga['unidade__sigla']] = []
                            unidade_vaga[vaga['unidade__sigla']].append(vaga)
                
                if unidade_vaga:
                    if inscrito['Unidade 1'] in unidade_vaga and unidade_vaga[inscrito['Unidade 1']]:
                        vaga = unidade_vaga[inscrito['Unidade 1']].pop(0)
                        vagas_rem.remove(vaga)
                        atribuicoes[classif.index(inscrito)] = vagas.index(vaga)

                    elif inscrito['Unidade 2'] in unidade_vaga and unidade_vaga[inscrito['Unidade 2']]:
                        vaga = unidade_vaga[inscrito['Unidade 2']].pop(0)
                        vagas_rem.remove(vaga)
                        atribuicoes[classif.index(inscrito)] = vagas.index(vaga)

                    elif inscrito['Interesse em outra unidade']== 'SIM':
                        if inscrito['CH MÊS'] > 36:
                            vaga = unidade_vaga[list(unidade_vaga.keys())[0]].pop(0)
                            vagas_rem.remove(vaga)
                            atribuicoes[classif.index(inscrito)] = vagas.index(vaga)
                        else:
                            espera.append(classif.index(inscrito))
                
                if(len(vagas_rem) == len(espera)):
                    break

            for idx_ag in espera:
                atribuicoes[idx_ag] = vagas.index(vagas_rem.pop(0))

            wb = openpyxl.load_workbook(os.path.join(os.path.dirname(__file__),'excel_files','DistribuiçãoTemplate.xltx'))
            wb.template = False
            sheet = wb['Distribuição']

            #printa os agentes com vagas

            agente_print_keys = ['IMPEDIMENTO','CH MÊS','CH TRIMESTRE','Nº','Nome','Matrícula','Unidade Lot.',\
                'Jornada de trab.','Nº plantão', 'Unidade 1', 'Unidade 2', 'Interesse em outra unidade']
            row = 2
            for agente in inscritos:
                # print(type(agente))
                id_agente = classif.index(agente)
                if id_agente in atribuicoes:
                    
                    dados_agente_printar = {key : agente[key] for key in agente_print_keys}
                    
                    for col, val in enumerate(list(dados_agente_printar.values())+[str(row-1)]+list(vagas[atribuicoes[id_agente]].values()), start=1):
                        sheet.cell(row=row, column=col).value = val
                    row+=1
            
            #printa os sem vaga 
            for agente in inscritos:
                id_agente = classif.index(agente)
                if id_agente not in atribuicoes:
                    dados_agente_printar = {key : agente[key] for key in agente_print_keys}
                    for col, val in enumerate(dados_agente_printar.values(), start=1):
                        sheet.cell(row=row, column=col).value = val
                    row+=1

            aba_confirmados = wb['Confirmados'] 
            aba_confirmados.cell(row=2, column=1).value = f"Prestação de serviço voluntário {dia.strftime('%d/%m')} (confirmados)"       
            aba_confirmados.cell(row=4, column=1).value = f"Trimestre = {' + '.join(trimestre)}" 

            response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/ms-excel')
            response['Content-Disposition'] = f'attachment; filename=Distribuição {dia.strftime(f"%d-%m")}.xlsx'
            return response


    return render(request, 'dashboard/distribuicao.html', {'form': form})


def atesto_impedimentos(request):
    form = AtestoImpedimentoForm()
    if request.method == 'POST':
        form = AtestoImpedimentoForm(request.POST, request.FILES)    
        if form.is_valid():
            file = form.cleaned_data.get('file')
            inicio = form.cleaned_data.get('inicio').replace(tzinfo=None)
            fim = form.cleaned_data.get('fim').replace(tzinfo=None)           

            book = openpyxl.load_workbook(filename=BytesIO(file.read()))


            header, voluntarios = extract_list_from_sheet(book['Voluntário'], header_row=2)

            required_in_header = ["Classificação","Servidor","Matrícula", "Unidade Lot.","DATA SV", "UNIDADE", "ATIVIDADE", "RESPONSÁVEL", "HORÁRIO", "CH"]
            if not all(x in header for x in required_in_header):
                messages.warning(request,f"Houve um problema com o cabeçalho da sua planilha, verifique se as informações estão corretas.\
                    \nObs: o cabeçalho deve conter EXATAMENTE estes textos: '"+"', '".join(required_in_header)+"'")
                return render(request, 'dashboard/atesto_impedimentos.html', {'form': form})

            vol_periodo = []
            for v in voluntarios:
                if v['DATA SV'] != None:
                    # if type(v['DATA SV'])== str:
                    #     v['DATA SV'] = datetime.strptime(v['DATA SV'], "%d-%b")
                    #     v['DATA SV'] = v['DATA SV'].replace(year=2020)
                    if v["DATA SV"] >= inicio and v["DATA SV"] <= fim:
                        vol_periodo.append(v)
            
            if not vol_periodo:
                messages.warning(request,f"A planilha enviada não possui voluntários no período solicitado!")
                return render(request, 'dashboard/atesto_impedimentos.html', {'form': form})


            por_unidade = {}
            for v in vol_periodo:
                if v['Unidade Lot.'] not in por_unidade:
                    por_unidade[v['Unidade Lot.']] = []
                por_unidade[v['Unidade Lot.']].append(v)

            
            mem_file = BytesIO()
            with zipfile.ZipFile(mem_file, "w") as zip_file:
                for unidade in por_unidade.keys():
                    vol_unidade = por_unidade[unidade]
                    wb = openpyxl.load_workbook(os.path.join(os.path.dirname(__file__),'excel_files', 'atesto_imp_template.xltx'))
                    wb.template = False
                    aba = wb.active
                    sb = NewSheetBuilder(aba)
                    sb.enter()
                    sb.enter()
                    sb.enter()
                    for vol in vol_unidade:
                        if vol['CH']:
                            sb.write_cell("")
                            sb.write_cell(vol['Servidor'])
                            sb.write_cell(vol['Matrícula'])
                            sb.write_cell(vol['Unidade Lot.'])
                            sb.write_cell(vol['Jornada \nde trab.'])
                            sb.write_cell(vol['Nº plantão'])
                            sb.write_cell(vol['DATA SV'].strftime("%d-%b"))
                            sb.write_cell(vol['UNIDADE'])
                            sb.write_cell(vol['HORÁRIO'])
                            sb.enter()
                    filename = f"{unidade} Atesto Impedimentos {inicio.strftime('%d-%m')} a {fim.strftime('%d-%m')}.xlsx"
                    
                    zip_file.writestr(filename, data=save_virtual_workbook(wb))
            
            mem_file.seek(0)
            filename = f"Atesto Impedimentos {inicio.strftime('%d-%m')} a {fim.strftime('%d-%m')}.zip"
            response = HttpResponse( mem_file, content_type='application/zip')
            response['Content-Disposition'] = 'attachment; filename='+filename
            return response
    context = {'form': form}
    return render(request, 'dashboard/atesto_impedimentos.html', context)


def homologacoes(request):
    form = HomologacaoForm()
    if request.method == 'POST':
        form = HomologacaoForm(request.POST, request.FILES)    
        if form.is_valid():
            file = form.cleaned_data.get('file')
            dia = form.cleaned_data.get('dia').replace(tzinfo=None)
                       

            book = openpyxl.load_workbook(filename=BytesIO(file.read()))


            header, voluntarios = extract_list_from_sheet(book['Voluntário'], header_row=2)
            
            required_in_header = ["Classificação","Servidor","Matrícula", "Unidade Lot.","DATA SV", "UNIDADE", "ATIVIDADE", "RESPONSÁVEL", "HORÁRIO", "CH"]
            if not all(x in header for x in required_in_header):
                messages.warning(request,f"Houve um problema com o cabeçalho da sua planilha, verifique se as informações estão corretas.\
                    \nObs: o cabeçalho deve conter EXATAMENTE estes textos: '"+"', '".join(required_in_header)+"'")
                return render(request, 'dashboard/homologações.html', {'form': form})

            vol_dia = []
            for v in voluntarios:
                if v['DATA SV'] != None:
                    if type(v['DATA SV'])== str:
                        v['DATA SV'] = datetime.strptime(v['DATA SV'], "%d-%b")
                        v['DATA SV'] = v['DATA SV'].replace(year=2020)
                    if v["DATA SV"] == dia:
                        vol_dia.append(v)
            
            if not vol_dia:
                messages.warning(request,f"A planilha enviada não possui voluntários no dia solicitado!")
                return render(request, 'dashboard/homologações.html', {'form': form})
            
            por_unidade = {}
            for v in vol_dia:
                if v['UNIDADE'] not in por_unidade:
                    por_unidade[v['UNIDADE']] = []
                por_unidade[v['UNIDADE']].append(v)


            
            mem_file = BytesIO()
            with zipfile.ZipFile(mem_file, "w") as zip_file:
                for unidade in por_unidade.keys():
                    vol_unidade = por_unidade[unidade]
                    wb = openpyxl.load_workbook(os.path.join(os.path.dirname(__file__),'excel_files', 'HomologaçãoTemplate.xltx'))
                    wb.template = False
                    sheet = wb.active
                    sheet.cell(row=2,column=2).value = f"{unidade}_Prestação de Serviço Voluntário"
                    
                    row = 4
                    vol_ordem_alpha = sorted(vol_unidade, key=lambda x: x['Servidor'], reverse=False)

                    for vol in vol_ordem_alpha:
                        if vol['CH']:
                            sheet.cell(row=row, column=2).value = vol['Classificação']
                            sheet.cell(row=row, column=3).value = vol['Servidor']
                            sheet.cell(row=row, column=4).value = vol['Matrícula']
                            sheet.cell(row=row, column=5).value = vol['Unidade Lot.']
                            sheet.cell(row=row, column=6).value = vol['DATA SV']
                            sheet.cell(row=row, column=7).value = vol['UNIDADE']
                            sheet.cell(row=row, column=8).value = vol['ATIVIDADE']
                            sheet.cell(row=row, column=9).value = vol['RESPONSÁVEL']
                            sheet.cell(row=row, column=10).value = vol['HORÁRIO']
                            sheet.cell(row=row, column=11).value = vol['CH']
                            row+=1

                    filename = f"{unidade} Homologação {dia.strftime('%d-%m')}.xlsx"
                    
                    zip_file.writestr(filename, data=save_virtual_workbook(wb))
                    
                    wb = openpyxl.load_workbook(os.path.join(os.path.dirname(__file__),'excel_files', 'HomologaçãoImprimirTemplate.xltx'))
                    wb.template = False
                    sheet = wb.active
                    sheet.cell(row=9,column=3).value = f"{dia.strftime('%d/%m/%Y')}"
                    sheet.cell(row=10,column=3).value = f"{unidade}"

                    

                    row = 14
                    for vol in vol_ordem_alpha:
                        if vol['CH']:
                            sheet.cell(row=row, column=2).value = vol['Matrícula']
                            sheet.cell(row=row, column=3).value = vol['Servidor']
                            row+=1
                       


                    filename = f"{unidade} Homologação {dia.strftime('%d-%m')} IMPRIMIR.xlsx"
                    
                    zip_file.writestr(filename, data=save_virtual_workbook(wb))
            
            mem_file.seek(0)
            filename = f"Homologações {dia.strftime('%d-%m')}.zip"
            response = HttpResponse( mem_file, content_type='application/zip')
            response['Content-Disposition'] = 'attachment; filename='+filename
            return response
    context = {'form': form}
    return render(request, 'dashboard/homologações.html', context)